"""
export_import.py — Экспорт медиатеки в JSON и Excel.

Структура:
  export_to_json()   → полный дамп всех данных (категории + записи по типам)
  export_to_excel()  → красиво отформатированная книга Excel:
      • Лист «Сводка»        — KPI и статистика по категориям
      • Лист «Аниме»         — все записи аниме
      • Лист «Манга»         — все записи манги
      • Лист «Фильмы»        — все фильмы
      • Лист «Сериалы»       — все сериалы
      • Лист «Книги»         — все книги
      • Лист «Игры»          — все игры

Импорт удалён намеренно — данные вносятся только через интерфейс.
"""

import json
import sqlite3
import os
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Путь к базе данных (относительно корня приложения) ──────────────────────
DATABASE = 'database.db'


# ════════════════════════════════════════════════════════════════════════════
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ════════════════════════════════════════════════════════════════════════════

def _get_conn():
    """Открывает соединение с БД в режиме только-для-чтения (row_factory)."""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def _calc_hours(item_type, data):
    """Вычисляет количество часов по тем же правилам, что и app.py."""
    try:
        if item_type == 'anime':
            mpl = 20 if data['anime_type'] == 'сериал' else 6
            ep  = int(data['episodes'] or 0)
            rw  = int(data['rewatches'] or 0)
            we  = int(data.get('watched_episodes') or 0)
            total_ep = ep * (1 + rw) + we
            return round(total_ep * mpl / 60, 1)

        elif item_type == 'manga':
            ch = int(data['chapters'] or 0)
            rr = int(data['rerereads'] or 0)
            rc = int(data.get('read_chapters') or 0)
            total_ch = ch * (1 + rr) + rc
            return round(total_ch * 5 / 60, 1)

        elif item_type == 'films':
            dur = int(data['duration'] or 0)
            rw  = int(data['rewatches'] or 0)
            return round(dur * (1 + rw) / 60, 1)

        elif item_type == 'series':
            ep  = int(data['episodes'] or 0)
            epd = int(data['episode_duration'] or 45)
            rw  = int(data['rewatches'] or 0)
            we  = int(data.get('watched_episodes') or 0)
            total_ep = ep * (1 + rw) + we
            return round(total_ep * epd / 60, 1)

        elif item_type == 'books':
            if data.get('book_type') == 'аудиокнига':
                return round(float(data.get('hours_reading') or 0), 1)
            pg = int(data.get('pages_duration') or 0)
            rr = int(data.get('rerereads') or 0)
            pr = int(data.get('pages_read') or 0)
            total_pg = pg * (1 + rr) + pr
            return round(total_pg / 60, 1)   # ~1 мин/стр

        elif item_type == 'games':
            return round(float(data.get('hours') or 0), 1)

    except Exception:
        pass
    return 0.0


def _status_ru(status):
    """Капитализирует и делает статус читаемым."""
    return (status or '').capitalize()


# ════════════════════════════════════════════════════════════════════════════
# JSON ЭКСПОРТ
# ════════════════════════════════════════════════════════════════════════════

def export_to_json():
    """
    Возвращает (json_string, filename) — полный дамп медиатеки.
    Включает категории и все записи с данными по каждому типу.
    """
    conn = _get_conn()
    cur  = conn.cursor()

    data = {
        'meta': {
            'app':         'Booek Media Tracker',
            'version':     '2.0',
            'export_date': datetime.now().isoformat(),
        },
        'categories': [],
        'items':      [],
    }

    # ── Категории ────────────────────────────────────────────────────────
    cur.execute('SELECT * FROM categories ORDER BY name')
    data['categories'] = [dict(r) for r in cur.fetchall()]

    # ── Базовые записи + специфические поля ─────────────────────────────
    cur.execute('SELECT * FROM items_base ORDER BY id')
    bases = cur.fetchall()

    # Подготовим словарь тип-таблица для быстрого поиска
    type_tables = {
        'anime':  'items_anime',
        'manga':  'items_manga',
        'films':  'items_films',
        'series': 'items_series',
        'books':  'items_books',
        'games':  'items_games',
    }

    # Словарь category_id → type
    cur.execute('SELECT id, type FROM categories')
    cat_types = {r['id']: r['type'] for r in cur.fetchall()}

    for base in bases:
        item_dict = dict(base)
        cat_type  = cat_types.get(base['category_id'])
        item_dict['category_type'] = cat_type

        # Специфические данные
        if cat_type and cat_type in type_tables:
            tbl = type_tables[cat_type]
            cur.execute(f'SELECT * FROM {tbl} WHERE item_id = ?', (base['id'],))
            specific = cur.fetchone()
            item_dict['type_data'] = dict(specific) if specific else {}
            item_dict['hours']     = _calc_hours(cat_type, item_dict['type_data'])
        else:
            item_dict['type_data'] = {}
            item_dict['hours']     = 0

        data['items'].append(item_dict)

    conn.close()

    json_str = json.dumps(data, ensure_ascii=False, indent=2)
    filename  = f"booek_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    return json_str, filename


# ════════════════════════════════════════════════════════════════════════════
# EXCEL ЭКСПОРТ — СТИЛИ
# ════════════════════════════════════════════════════════════════════════════

# Цветовая палитра (hex без #)
PALETTE = {
    'header_bg':   '1C1C1E',   # тёмный фон заголовка
    'header_fg':   'FFFFFF',   # белый текст заголовка
    'accent':      '0A84FF',   # синий акцент
    'accent_light':'D6EAFF',   # светло-синий
    'done':        '34C759',   # зелёный — завершено
    'done_light':  'D4EDDA',
    'progress':    '0A84FF',   # синий — в процессе
    'progress_light': 'D6EAFF',
    'planned':     'FF9F0A',   # жёлтый — планирую
    'planned_light':'FFF3CD',
    'dropped':     'FF453A',   # красный — брошено
    'dropped_light':'FDDCDA',
    'stripe':      'F5F5F7',   # чередование строк
    'white':       'FFFFFF',
    'border':      'D1D1D6',
}

DONE_STATUSES     = {'просмотрено', 'прочитано', 'прошёл целиком'}
PROGRESS_STATUSES = {'смотрю', 'читаю', 'играю'}
PLANNED_STATUSES  = {'планирую'}
DROPPED_STATUSES  = {'не досмотрел', 'не дочитал', 'прошёл частично'}


def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, color='000000', size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name='Calibri')


def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _thin_border():
    s = Side(style='thin', color=PALETTE['border'])
    return Border(left=s, right=s, top=s, bottom=s)


def _status_fill(status):
    """Возвращает заливку строки по статусу."""
    s = (status or '').lower()
    if s in DONE_STATUSES:
        return _fill(PALETTE['done_light'])
    if s in PROGRESS_STATUSES:
        return _fill(PALETTE['progress_light'])
    if s in PLANNED_STATUSES:
        return _fill(PALETTE['planned_light'])
    if s in DROPPED_STATUSES:
        return _fill(PALETTE['dropped_light'])
    return None


def _write_header_row(ws, headers, row=1, bg=PALETTE['header_bg'], fg=PALETTE['header_fg']):
    """Записывает строку заголовков с форматированием."""
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.fill   = _fill(bg)
        cell.font   = _font(bold=True, color=fg, size=11)
        cell.alignment = _align('center', 'center')
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 22


def _style_data_row(ws, row_idx, n_cols, status=None, stripe=False):
    """Применяет форматирование к строке данных."""
    fill = _status_fill(status)
    if fill is None:
        fill = _fill(PALETTE['stripe']) if stripe else _fill(PALETTE['white'])
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill      = fill
        cell.border    = _thin_border()
        cell.alignment = _align('left', 'center', wrap=False)
    ws.row_dimensions[row_idx].height = 18


# ════════════════════════════════════════════════════════════════════════════
# EXCEL ЭКСПОРТ — ЛИСТ «СВОДКА»
# ════════════════════════════════════════════════════════════════════════════

def _build_summary_sheet(ws, conn):
    """Лист 1 — Сводка: общие KPI + таблица по категориям."""
    cur = conn.cursor()

    # ── Заголовок листа ──────────────────────────────────────────────────
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value     = '📚 Booek — Сводка медиатеки'
    title_cell.fill      = _fill(PALETTE['header_bg'])
    title_cell.font      = _font(bold=True, color=PALETTE['header_fg'], size=16)
    title_cell.alignment = _align('center', 'center')
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:G2')
    dt_cell = ws['A2']
    dt_cell.value     = f"Экспортировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    dt_cell.font      = _font(italic=True, color='6E6E73', size=10)
    dt_cell.alignment = _align('center', 'center')
    ws.row_dimensions[2].height = 16

    # ── Таблица по категориям ─────────────────────────────────────────────
    cur.execute('SELECT * FROM categories ORDER BY name')
    categories = cur.fetchall()

    _write_header_row(ws, ['Категория', 'Тип', 'Записей', 'Часов', 'Завершено', 'В процессе', 'В планах'],
                      row=4)

    DONE_S     = ['просмотрено', 'прочитано', 'прошёл целиком']
    PROGRESS_S = ['смотрю', 'читаю', 'играю']
    PLANNED_S  = ['планирую']
    TYPE_MAP   = {
        'anime': 'items_anime', 'manga': 'items_manga', 'films': 'items_films',
        'series': 'items_series', 'books': 'items_books', 'games': 'items_games',
    }
    TYPE_RU = {
        'anime': 'Аниме', 'manga': 'Манга', 'films': 'Фильмы',
        'series': 'Сериалы', 'books': 'Книги', 'games': 'Игры',
    }

    total_items = total_hours = total_done = total_prog = total_plan = 0

    for row_i, cat in enumerate(categories, start=5):
        cat_type = cat['type']
        tbl      = TYPE_MAP.get(cat_type)
        cat_name = cat['name']

        count = done = prog = plan = 0
        hours = 0.0

        if tbl:
            cur.execute(f'SELECT COUNT(*) FROM {tbl} WHERE item_id IN (SELECT id FROM items_base WHERE category_id = ?)', (cat['id'],))
            count = cur.fetchone()[0]

            # Статусы
            ph = ','.join('?' * len(DONE_S))
            cur.execute(f'SELECT COUNT(*) FROM {tbl} WHERE status IN ({ph}) AND item_id IN (SELECT id FROM items_base WHERE category_id = ?)', DONE_S + [cat['id']])
            done = cur.fetchone()[0]
            cur.execute(f'SELECT COUNT(*) FROM {tbl} WHERE status IN ({",".join("?" * len(PROGRESS_S))}) AND item_id IN (SELECT id FROM items_base WHERE category_id = ?)', PROGRESS_S + [cat['id']])
            prog = cur.fetchone()[0]
            cur.execute(f'SELECT COUNT(*) FROM {tbl} WHERE status IN ({",".join("?" * len(PLANNED_S))}) AND item_id IN (SELECT id FROM items_base WHERE category_id = ?)', PLANNED_S + [cat['id']])
            plan = cur.fetchone()[0]

            # Часы через Python (переиспользуем _calc_hours)
            cur.execute(f'SELECT * FROM {tbl} t JOIN items_base b ON t.item_id = b.id WHERE b.category_id = ?', (cat['id'],))
            for r in cur.fetchall():
                hours += _calc_hours(cat_type, dict(r))

        hours = round(hours, 1)
        total_items += count; total_hours += hours
        total_done  += done;  total_prog  += prog; total_plan += plan

        stripe = (row_i % 2 == 0)
        vals = [cat_name, TYPE_RU.get(cat_type, cat_type), count, hours, done, prog, plan]
        for col_i, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill      = _fill(PALETTE['stripe']) if stripe else _fill(PALETTE['white'])
            cell.border    = _thin_border()
            cell.alignment = _align('center' if col_i > 2 else 'left', 'center')
            cell.font      = _font(size=11)
        ws.row_dimensions[row_i].height = 18

    # Итого
    total_row = len(categories) + 5
    for col_i, val in enumerate(['ИТОГО', '', total_items, round(total_hours, 1), total_done, total_prog, total_plan], start=1):
        cell = ws.cell(row=total_row, column=col_i, value=val)
        cell.fill      = _fill(PALETTE['accent_light'])
        cell.font      = _font(bold=True, size=11)
        cell.border    = _thin_border()
        cell.alignment = _align('center' if col_i > 2 else 'left', 'center')
    ws.row_dimensions[total_row].height = 20

    # Ширины
    for col, w in zip('ABCDEFG', [28, 12, 10, 10, 12, 12, 10]):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A5'


# ════════════════════════════════════════════════════════════════════════════
# EXCEL ЭКСПОРТ — КАТЕГОРИЙНЫЕ ЛИСТЫ
# ════════════════════════════════════════════════════════════════════════════

def _build_category_sheet(ws, conn, cat, cat_type):
    """Строит лист для одной категории."""
    cur = conn.cursor()

    TYPE_MAP = {
        'anime': 'items_anime', 'manga': 'items_manga', 'films': 'items_films',
        'series': 'items_series', 'books': 'items_books', 'games': 'items_games',
    }
    tbl = TYPE_MAP.get(cat_type)
    if not tbl:
        return

    # Получаем записи
    cur.execute(
        f'SELECT b.*, t.* FROM items_base b JOIN {tbl} t ON b.id = t.item_id '
        f'WHERE b.category_id = ? ORDER BY b.title',
        (cat['id'],)
    )
    rows = cur.fetchall()
    if not rows:
        ws.cell(row=1, column=1, value='Нет записей').font = _font(italic=True, color='999999')
        return

    # ── Заголовок листа ──────────────────────────────────────────────────
    cat_name = cat['name']
    headers_map = {
        'anime':  ['№', 'Название', 'Тип аниме', 'Серии', 'Просм.серий', 'Пересм.', 'Статус', 'Часов', 'Источник'],
        'manga':  ['№', 'Название', 'Тип манги', 'Главы', 'Прочитано гл.', 'Перечит.', 'Статус', 'Часов', 'Источник'],
        'films':  ['№', 'Название', 'Длит. (мин)', 'Пересм.', 'Статус', 'Часов', 'Источник'],
        'series': ['№', 'Название', 'Серий', 'Мин/серия', 'Просм.серий', 'Пересм.', 'Статус', 'Часов', 'Источник'],
        'books':  ['№', 'Название', 'Тип книги', 'Страниц', 'Прочитано стр.', 'Часов (аудио)', 'Перечит.', 'Статус', 'Часов', 'Источник'],
        'games':  ['№', 'Название', 'Статус', 'Часов', 'Источник'],
    }
    headers = headers_map.get(cat_type, ['№', 'Название', 'Статус', 'Часов'])

    # Мерж заголовочной строки
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    h_cell = ws['A1']
    h_cell.value     = cat_name
    h_cell.fill      = _fill(PALETTE['header_bg'])
    h_cell.font      = _font(bold=True, color=PALETTE['header_fg'], size=14)
    h_cell.alignment = _align('center', 'center')
    ws.row_dimensions[1].height = 28

    _write_header_row(ws, headers, row=2)
    ws.freeze_panes = 'A3'

    # ── Строки данных ─────────────────────────────────────────────────────
    for row_i, r in enumerate(rows, start=3):
        d      = dict(r)
        status = d.get('status', '')
        hours  = _calc_hours(cat_type, d)
        idx    = row_i - 2

        if cat_type == 'anime':
            vals = [idx, d.get('title'), d.get('anime_type'), d.get('episodes'),
                    d.get('watched_episodes') or 0, d.get('rewatches') or 0,
                    _status_ru(status), hours, d.get('source_url') or '']

        elif cat_type == 'manga':
            vals = [idx, d.get('title'), d.get('manga_type'), d.get('chapters'),
                    d.get('read_chapters') or 0, d.get('rerereads') or 0,
                    _status_ru(status), hours, d.get('source_url') or '']

        elif cat_type == 'films':
            vals = [idx, d.get('title'), d.get('duration'),
                    d.get('rewatches') or 0, _status_ru(status), hours,
                    d.get('source_url') or '']

        elif cat_type == 'series':
            vals = [idx, d.get('title'), d.get('episodes'), d.get('episode_duration'),
                    d.get('watched_episodes') or 0, d.get('rewatches') or 0,
                    _status_ru(status), hours, d.get('source_url') or '']

        elif cat_type == 'books':
            vals = [idx, d.get('title'), d.get('book_type'), d.get('pages_duration'),
                    d.get('pages_read') or 0, d.get('hours_reading') or 0,
                    d.get('rerereads') or 0, _status_ru(status), hours,
                    d.get('source_url') or '']

        elif cat_type == 'games':
            vals = [idx, d.get('title'), _status_ru(status), hours,
                    d.get('source_url') or '']

        else:
            vals = [idx, d.get('title'), _status_ru(status), hours]

        stripe = (row_i % 2 == 0)
        for col_i, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.font      = _font(size=10)
            cell.border    = _thin_border()
            cell.alignment = _align('left' if col_i == 2 else 'center', 'center')

        _style_data_row(ws, row_i, len(headers), status=status, stripe=stripe)
        # Название — всегда слева и жирнее
        ws.cell(row=row_i, column=2).font = _font(bold=True, size=10)

    # ── Итого часов ───────────────────────────────────────────────────────
    total_row = len(rows) + 3
    total_hours = sum(_calc_hours(cat_type, dict(r)) for r in rows)
    ws.cell(row=total_row, column=1, value='Итого:').font = _font(bold=True)
    ws.cell(row=total_row, column=1).fill = _fill(PALETTE['accent_light'])
    # Ставим сумму в колонку «Часов»
    hours_col = len(headers) - 1  # предпоследняя колонка — всегда «Часов»
    ws.cell(row=total_row, column=hours_col, value=round(total_hours, 1)).font = _font(bold=True)
    ws.cell(row=total_row, column=hours_col).fill = _fill(PALETTE['accent_light'])

    # ── Ширины колонок ────────────────────────────────────────────────────
    col_widths_map = {
        'anime':  [5, 32, 14, 8, 12, 8, 16, 8, 40],
        'manga':  [5, 32, 14, 8, 14, 8, 16, 8, 40],
        'films':  [5, 32, 12, 8, 16, 8, 40],
        'series': [5, 32, 8, 10, 12, 8, 16, 8, 40],
        'books':  [5, 32, 14, 10, 14, 12, 8, 16, 8, 40],
        'games':  [5, 32, 16, 8, 40],
    }
    widths = col_widths_map.get(cat_type, [5, 32, 16, 8, 40])
    for col_i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = w


# ════════════════════════════════════════════════════════════════════════════
# EXCEL ЭКСПОРТ — ТОЧКА ВХОДА
# ════════════════════════════════════════════════════════════════════════════

def export_to_excel():
    """
    Возвращает (BytesIO, filename).
    Книга содержит лист Сводка + по одному листу на каждую категорию.
    """
    conn = _get_conn()
    cur  = conn.cursor()

    wb = openpyxl.Workbook()
    # Удаляем пустой лист по умолчанию
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # ── Лист 1: Сводка ───────────────────────────────────────────────────
    ws_summary = wb.create_sheet('📊 Сводка')
    _build_summary_sheet(ws_summary, conn)

    # ── Листы по категориям ───────────────────────────────────────────────
    SHEET_NAMES = {
        'anime': '✨ Аниме', 'manga': '📖 Манга', 'films': '🎬 Фильмы',
        'series': '📺 Сериалы', 'books': '📚 Книги', 'games': '🎮 Игры',
    }
    # Гарантированный порядок
    ORDER = ['anime', 'manga', 'films', 'series', 'books', 'games']

    cur.execute('SELECT * FROM categories ORDER BY name')
    categories = {r['type']: r for r in cur.fetchall()}

    for cat_type in ORDER:
        cat = categories.get(cat_type)
        if not cat:
            continue
        sheet_name = SHEET_NAMES.get(cat_type, cat['name'])[:31]  # Excel limit
        ws = wb.create_sheet(sheet_name)
        _build_category_sheet(ws, conn, cat, cat_type)

    conn.close()

    # ── Сохраняем в BytesIO (не нужен диск) ─────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"booek_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buf, filename
